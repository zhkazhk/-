from flask import Flask, request, jsonify, render_template, send_file
import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
import sqlite3
import threading
import time
import sys
import webbrowser
from flask_cors import CORS

# ========================= 核心配置与全局变量 =========================
# 全局配置
DEFAULT_CONFIG = {
    "black_overprint_price": 0.06,
    "color_overprint_price": 0.6,
    "default_period": "2026.01.01-2026.02.28"
}

# 解决打包后路径问题（统一所有路径使用此函数）
def resource_path(relative_path):
    """获取打包后文件的绝对路径（全局统一使用）"""
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller 创建的临时文件夹
        base_path = sys._MEIPASS
    else:
        # 开发环境路径
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# 全局标记：避免重复初始化数据库
DB_INITIALIZED = False

# ========================= 数据库操作（修复路径问题） =========================
def init_db():
    """初始化数据库（统一使用resource_path）"""
    global DB_INITIALIZED
    if DB_INITIALIZED:
        return
    
    # 统一使用resource_path获取数据库路径
    db_path = resource_path("dist/printer_fee.db")
    # 确保dist文件夹存在
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    
    # 建立连接，关闭线程检查（关键）
    conn = sqlite3.connect(db_path, check_same_thread=False)
    cursor = conn.cursor()
    
    try:
        # 创建客户表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name TEXT UNIQUE NOT NULL,
            invoice_type TEXT DEFAULT '增税',
            create_time TEXT,
            update_time TEXT
        )
        ''')
        
        # 创建计算历史表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS calculation_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER,
            company_name TEXT,
            invoice_type TEXT,
            location TEXT,
            ip TEXT,
            model TEXT,
            serial TEXT,
            first_date TEXT,
            second_date TEXT,
            first_black INTEGER,
            first_color INTEGER,
            second_black INTEGER,
            second_color INTEGER,
            package_black INTEGER,
            package_color INTEGER,
            basic_fee REAL,
            used_black INTEGER,
            used_color INTEGER,
            over_black INTEGER,
            over_color INTEGER,
            over_fee_black REAL,
            over_fee_color REAL,
            total_fee REAL,
            period TEXT,
            black_price REAL,
            color_price REAL,
            calculate_time TEXT,
            FOREIGN KEY (customer_id) REFERENCES customers (id)
        )
        ''')
        
        conn.commit()
        DB_INITIALIZED = True
        print(f"数据库初始化成功！路径：{db_path}")
    except Exception as e:
        conn.rollback()
        print(f"数据库初始化失败：{str(e)}")
        raise
    finally:
        try:
            if not conn.closed:
                conn.close()
        except:
            pass

def add_or_update_customer(company_name, invoice_type):
    """添加/更新客户信息（统一路径）"""
    db_path = resource_path("dist/printer_fee.db")
    conn = sqlite3.connect(db_path, check_same_thread=False)
    cursor = conn.cursor()
    
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute('SELECT id FROM customers WHERE company_name = ?', (company_name,))
        result = cursor.fetchone()
        
        if result:
            customer_id = result[0]
            cursor.execute('''
            UPDATE customers SET invoice_type = ?, update_time = ? WHERE id = ?
            ''', (invoice_type, now, customer_id))
        else:
            cursor.execute('''
            INSERT INTO customers (company_name, invoice_type, create_time, update_time)
            VALUES (?, ?, ?, ?)
            ''', (company_name, invoice_type, now, now))
            customer_id = cursor.lastrowid
        
        conn.commit()
        return customer_id
    except Exception as e:
        conn.rollback()
        print(f"添加/更新客户失败：{str(e)}")
        raise
    finally:
        try:
            conn.close()
        except:
            pass

def get_customer_list():
    """获取所有客户名称列表"""
    db_path = resource_path("dist/printer_fee.db")
    conn = sqlite3.connect(db_path, check_same_thread=False)
    cursor = conn.cursor()
    
    try:
        cursor.execute('SELECT company_name FROM customers ORDER BY update_time DESC')
        rows = cursor.fetchall()
        return [row[0] for row in rows]
    except Exception as e:
        print(f"获取客户列表失败：{str(e)}")
        return []
    finally:
        try:
            conn.close()
        except:
            pass

def get_customer_info(company_name):
    """根据公司名称获取客户详情"""
    db_path = resource_path("dist/printer_fee.db")
    conn = sqlite3.connect(db_path, check_same_thread=False)
    cursor = conn.cursor()
    
    try:
        cursor.execute('SELECT * FROM customers WHERE company_name = ?', (company_name,))
        row = cursor.fetchone()
        if row:
            return {
                "id": row[0],
                "company_name": row[1],
                "invoice_type": row[2]
            }
        return None
    except Exception as e:
        print(f"获取客户详情失败：{str(e)}")
        return None
    finally:
        try:
            conn.close()
        except:
            pass

def get_last_meter_data(company_name, model, serial):
    """获取最后一次抄表数据"""
    db_path = resource_path("dist/printer_fee.db")
    conn = sqlite3.connect(db_path, check_same_thread=False)
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
        SELECT second_black, second_color, second_date 
        FROM calculation_history 
        WHERE company_name = ? AND model = ? AND serial = ?
        ORDER BY calculate_time DESC LIMIT 1
        ''', (company_name, model, serial))
        row = cursor.fetchone()
        if row:
            return {
                "last_black": row[0],
                "last_color": row[1],
                "last_date": row[2]
            }
        return None
    except Exception as e:
        print(f"获取抄表数据失败：{str(e)}")
        return None
    finally:
        try:
            conn.close()
        except:
            pass

def add_calculation(data):
    """添加计算记录"""
    db_path = resource_path("dist/printer_fee.db")
    conn = sqlite3.connect(db_path, check_same_thread=False)
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
        INSERT INTO calculation_history (
            customer_id, company_name, invoice_type,
            location, ip, model, serial, first_date, second_date,
            first_black, first_color, second_black, second_color,
            package_black, package_color, basic_fee, used_black, used_color,
            over_black, over_color, over_fee_black, over_fee_color,
            total_fee, period, black_price, color_price, calculate_time
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('customer_id', 0),
            data.get('company_name', ''),
            data.get('invoice_type', '增税'),
            data['location'], data['ip'], data['model'], data['serial'],
            data['first_date'], data['second_date'], data['first_black'],
            data['first_color'], data['second_black'], data['second_color'],
            data['package_black'], data['package_color'], data['basic_fee'],
            data['used_black'], data['used_color'], data['over_black'],
            data['over_color'], data['over_fee_black'], data['over_fee_color'],
            data['total_fee'], data['period'], data['black_price'],
            data['color_price'], data['calculate_time']
        ))
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"添加计算记录失败：{str(e)}")
        raise
    finally:
        try:
            conn.close()
        except:
            pass

def get_all_calculations():
    """获取所有计算记录"""
    db_path = resource_path("dist/printer_fee.db")
    conn = sqlite3.connect(db_path, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    try:
        cursor.execute('SELECT * FROM calculation_history ORDER BY calculate_time DESC')
        rows = cursor.fetchall()
        return [dict(row) for row in rows]
    except Exception as e:
        print(f"获取计算记录失败：{str(e)}")
        return []
    finally:
        try:
            conn.close()
        except:
            pass

def get_customer_calculations(company_name):
    """获取指定客户的所有计算记录"""
    db_path = resource_path("dist/printer_fee.db")
    conn = sqlite3.connect(db_path, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
        SELECT * FROM calculation_history 
        WHERE company_name = ? 
        ORDER BY calculate_time DESC
        ''', (company_name,))
        rows = cursor.fetchall()
        return [dict(row) for row in rows]
    except Exception as e:
        print(f"获取客户计算记录失败：{str(e)}")
        return []
    finally:
        try:
            conn.close()
        except:
            pass

def clear_calculations():
    """清空计算记录"""
    db_path = resource_path("dist/printer_fee.db")
    conn = sqlite3.connect(db_path, check_same_thread=False)
    cursor = conn.cursor()
    
    try:
        cursor.execute('DELETE FROM calculation_history')
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"清空计算记录失败：{str(e)}")
        raise
    finally:
        try:
            conn.close()
        except:
            pass

# ========================= 核心业务逻辑（无修改） =========================
def validate_inputs(data):
    """验证输入数据有效性"""
    try:
        if not data.get("company_name", "").strip():
            return {
                "valid": False,
                "error": "公司名称不能为空"
            }
        
        numeric_fields = [
            "first_black", "first_color", "second_black", "second_color",
            "package_black", "package_color", "basic_fee"
        ]
        for field in numeric_fields:
            float(data.get(field, 0))
        
        first_black = int(data.get("first_black", 0))
        second_black = int(data.get("second_black", 0))
        first_color = int(data.get("first_color", 0))
        second_color = int(data.get("second_color", 0))
        
        warnings = []
        if second_black < first_black:
            warnings.append("第二次抄表黑色张数不能小于第一次")
        if second_color < first_color:
            warnings.append("第二次抄表彩色张数不能小于第一次")
        
        return {
            "valid": True,
            "warnings": warnings
        }
    except ValueError as e:
        return {
            "valid": False,
            "error": f"数字输入错误：{str(e)}"
        }

def calculate_cost(data, config=DEFAULT_CONFIG):
    """核心计算逻辑"""
    try:
        company_name = data.get("company_name", "")
        invoice_type = data.get("invoice_type", "增税")
        
        location = data.get("location", "")
        ip = data.get("ip", "")
        model = data.get("model", "")
        serial = data.get("serial", "")
        
        first_date = data.get("first_date", "2025.10.31")
        second_date = data.get("second_date", "2025.12.31")
        first_black = int(data.get("first_black", 0))
        first_color = int(data.get("first_color", 0))
        second_black = int(data.get("second_black", 0))
        second_color = int(data.get("second_color", 0))
        
        package_black = int(data.get("package_black", 0))
        package_color = int(data.get("package_color", 0))
        basic_fee = float(data.get("basic_fee", 0))
        
        black_price = float(data.get("black_price", config["black_overprint_price"]))
        color_price = float(data.get("color_price", config["color_overprint_price"]))
        period = data.get("period", config["default_period"])
        
        # 核心计算
        used_black = second_black - first_black
        used_color = second_color - first_color
        
        over_black = max(used_black - package_black, 0)
        over_color = max(used_color - package_color, 0)
        
        over_fee_black = round(over_black * black_price, 2)
        over_fee_color = round(over_color * color_price, 2)
        
        total_fee = round(over_fee_black + over_fee_color + basic_fee, 2)
        
        # 组装结果
        result = {
            "customer_id": data.get("customer_id", 0),
            "company_name": company_name,
            "invoice_type": invoice_type,
            "location": location,
            "ip": ip,
            "model": model,
            "serial": serial,
            "first_date": first_date,
            "second_date": second_date,
            "first_black": first_black,
            "first_color": first_color,
            "second_black": second_black,
            "second_color": second_color,
            "package_black": package_black,
            "package_color": package_color,
            "basic_fee": basic_fee,
            "used_black": used_black,
            "used_color": used_color,
            "over_black": over_black,
            "over_color": over_color,
            "over_fee_black": over_fee_black,
            "over_fee_color": over_fee_color,
            "total_fee": total_fee,
            "period": period,
            "black_price": black_price,
            "color_price": color_price,
            "calculate_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        return {
            "success": True,
            "data": result,
            "warnings": []
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

def export_to_excel(config=DEFAULT_CONFIG):
    """Excel导出逻辑"""
    try:
        data_list = get_all_calculations()
        if not data_list:
            return {"success": False, "error": "暂无计算数据可导出"}
        
        wb = Workbook()
        ws = wb.active
        ws.title = "打印机费用清单"
        
        # 样式定义
        title_font = Font(name="微软雅黑", size=14, bold=True)
        header_font = Font(name="微软雅黑", size=10, bold=True)
        normal_font = Font(name="微软雅黑", size=9)
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        total_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        no_border = Border(
            left=Side(style='none'),
            right=Side(style='none'),
            top=Side(style='none'),
            bottom=Side(style='none')
        )
        
        # 1. 标题部分
        ws.merge_cells('A1:R1')
        ws['A1'] = "上海库克打印机有限公司开票清单"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment
        
        ws.merge_cells('A2:B2')
        ws['A2'] = f"客户名称：{data_list[0]['company_name']}"
        ws['A2'].font = Font(name="微软雅黑", size=11)
        ws['A2'].alignment = left_alignment
        
        ws.merge_cells('C2:D2')
        ws['C2'] = f"发票类型：{data_list[0]['invoice_type']}"
        ws['C2'].font = Font(name="微软雅黑", size=11)
        ws['C2'].alignment = left_alignment
        
        # 2. 列标题
        headers_row3 = [
            "客户名称", "机器位置", "IP地址", "设备型号", "设备序号",
            f"{data_list[0]['first_date']}初始张数", "",
            "基本费（元）", "包印张数", "",
            f"{data_list[0]['second_date']}抄表张数", "",
            "使用张数", "", "超张数", "", "", "", "", ""
        ]
        
        headers_row4 = [
            "", "", "", "", "", "黑色", "彩色", "", "黑色", "彩色",
            "黑色", "彩色", "黑色", "彩色", "黑色", "彩色", "黑色", "彩色", "超印费小计"
        ]
        
        # 填写第3行标题
        for col, header in enumerate(headers_row3, 1):
            if col > 19:
                break
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        # 合并单元格
        ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=7)
        ws.merge_cells(start_row=3, start_column=9, end_row=3, end_column=10)
        ws.merge_cells(start_row=3, start_column=11, end_row=3, end_column=12)
        ws.merge_cells(start_row=3, start_column=13, end_row=3, end_column=14)
        ws.merge_cells(start_row=3, start_column=15, end_row=3, end_column=16)
        ws.merge_cells(start_row=3, start_column=17, end_row=3, end_column=19)
        
        # 重新设置合并单元格样式
        for start_col in [6, 9, 11, 13, 15, 17]:
            merged_cell = ws.cell(row=3, column=start_col)
            merged_cell.font = header_font
            merged_cell.alignment = center_alignment
            merged_cell.fill = header_fill
            merged_cell.border = thin_border
        
        # 填写第4行标题
        for col, header in enumerate(headers_row4, 1):
            if col > 19:
                break
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        # 3. 填写数据行
        data_start_row = 5
        total_over_fee = 0.0
        total_basic_fee = 0.0
        total_all_fee = 0.0
        
        # 按客户名称分组
        customer_groups = {}
        for data in data_list:
            if data['company_name'] not in customer_groups:
                customer_groups[data['company_name']] = []
            customer_groups[data['company_name']].append(data)
        
        # 遍历分组填写数据
        current_row = data_start_row
        for company_name, group_data in customer_groups.items():
            # 客户分组标题
            ws.merge_cells(f'A{current_row}:R{current_row}')
            group_cell = ws.cell(row=current_row, column=1, value=f"【{company_name}】")
            group_cell.font = Font(name="微软雅黑", size=10, bold=True)
            group_cell.alignment = center_alignment
            group_cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            current_row += 1
            
            # 填写明细数据
            for data in group_data:
                basic_fee_display = f"{data['basic_fee']:.2f}"
                
                row_data = [
                    company_name, data['location'], data['ip'], data['model'], data['serial'],
                    data['first_black'], data['first_color'], basic_fee_display,
                    data['package_black'], data['package_color'],
                    data['second_black'], data['second_color'],
                    data['used_black'], data['used_color'],
                    data['over_black'], data['over_color'],
                    data['over_fee_black'], data['over_fee_color'],
                    data['over_fee_black'] + data['over_fee_color']
                ]
                
                for col, value in enumerate(row_data, 1):
                    if col > 19:
                        break
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.font = normal_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                total_over_fee += data['over_fee_black'] + data['over_fee_color']
                total_basic_fee += data['basic_fee']
                total_all_fee += data['total_fee']
                current_row += 1
        
        # 4. 汇总行
        summary_row = current_row + 3
        period = data_list[0]['period']
        
        # 租赁费汇总
        ws.merge_cells(f'A{summary_row}:B{summary_row}')
        ws[f'A{summary_row}'] = "租赁费"
        ws[f'A{summary_row}'].font = header_font
        ws[f'A{summary_row}'].alignment = center_alignment
        ws[f'A{summary_row}'].fill = header_fill
        ws[f'A{summary_row}'].border = no_border
        
        ws[f'C{summary_row}'] = period
        ws[f'C{summary_row}'].font = normal_font
        ws[f'C{summary_row}'].alignment = center_alignment
        ws[f'C{summary_row}'].border = no_border
        
        ws[f'E{summary_row}'] = f"¥{total_basic_fee:.2f}"
        ws[f'E{summary_row}'].font = normal_font
        ws[f'E{summary_row}'].alignment = center_alignment
        ws[f'E{summary_row}'].border = no_border
        
        # 超印费汇总
        summary_row2 = summary_row + 1
        ws.merge_cells(f'A{summary_row2}:B{summary_row2}')
        ws[f'A{summary_row2}'] = "超印费"
        ws[f'A{summary_row2}'].font = header_font
        ws[f'A{summary_row2}'].alignment = center_alignment
        ws[f'A{summary_row2}'].fill = header_fill
        ws[f'A{summary_row2}'].border = no_border
        
        ws[f'C{summary_row2}'] = f"{data_list[0]['first_date'][:7]}-{data_list[0]['second_date'][:7]}"
        ws[f'C{summary_row2}'].font = normal_font
        ws[f'C{summary_row2}'].alignment = center_alignment
        ws[f'C{summary_row2}'].border = no_border
        
        ws[f'E{summary_row2}'] = f"¥{total_over_fee:.2f}"
        ws[f'E{summary_row2}'].font = normal_font
        ws[f'E{summary_row2}'].alignment = center_alignment
        ws[f'E{summary_row2}'].border = no_border
        
        # 总费用汇总
        summary_row3 = summary_row2 + 1
        ws.merge_cells(f'A{summary_row3}:D{summary_row3}')
        ws[f'A{summary_row3}'] = "总费用"
        ws[f'A{summary_row3}'].font = Font(name="微软雅黑", size=12, bold=True)
        ws[f'A{summary_row3}'].alignment = center_alignment
        ws[f'A{summary_row3}'].fill = total_fill
        ws[f'A{summary_row3}'].border = no_border
        
        ws[f'E{summary_row3}'] = f"¥{total_all_fee:.2f}"
        ws[f'E{summary_row3}'].font = Font(name="微软雅黑", size=12, bold=True)
        ws[f'E{summary_row3}'].alignment = center_alignment
        ws[f'E{summary_row3}'].border = no_border
        ws[f'E{summary_row3}'].fill = total_fill
        
        # 调整列宽
        column_widths = [20, 25, 15, 18, 18, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 15]
        for col, width in enumerate(column_widths, 1):
            if col > 19:
                break
            ws.column_dimensions[chr(64 + col)].width = width
        
        # 调整行高
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 30
        ws.row_dimensions[4].height = 20
        
        # 保存文件（统一路径）
        filename = f"打印机费用清单_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        save_path = resource_path(filename)
        wb.save(save_path)
        
        return {
            "success": True,
            "filename": filename,
            "path": save_path,
            "total_basic_fee": total_basic_fee,
            "total_over_fee": total_over_fee,
            "total_all_fee": total_all_fee
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

# ========================= Flask应用初始化（仅初始化一次） =========================
# 仅创建一次Flask app（关键修复）
app = Flask(__name__, 
            static_folder=resource_path("static"),
            template_folder=resource_path("templates"))
CORS(app)  # 跨域支持

# ========================= Flask路由 =========================
@app.route('/')
def index():
    """首页"""
    customer_list = get_customer_list()
    return render_template('index.html', 
                           default_config=DEFAULT_CONFIG,
                           customer_list=customer_list)

@app.route('/api/get-customer-info', methods=['POST'])
def api_get_customer_info():
    """获取客户详情"""
    try:
        if not request.is_json:
            return jsonify({"success": False, "error": "请求格式必须为JSON"}), 400
        data = request.get_json()
        company_name = data.get('company_name', '').strip()
        if not company_name:
            return jsonify({"success": False, "error": "公司名称不能为空"}), 400
        
        customer_info = get_customer_info(company_name)
        if customer_info:
            return jsonify({"success": True, "data": customer_info})
        return jsonify({"success": False, "error": "客户信息不存在"})
    except Exception as e:
        return jsonify({"success": False, "error": f"服务器错误：{str(e)}"}), 500

@app.route('/api/get-last-meter-data', methods=['POST'])
def api_get_last_meter_data():
    """获取最后一次抄表数据"""
    try:
        if not request.is_json:
            return jsonify({"success": False, "error": "请求格式必须为JSON"}), 400
        data = request.get_json()
        company_name = data.get('company_name', '').strip()
        model = data.get('model', '').strip()
        serial = data.get('serial', '').strip()
        
        if not company_name or not model or not serial:
            return jsonify({"success": False, "error": "公司名称、型号、序号不能为空"}), 400
        
        meter_data = get_last_meter_data(company_name, model, serial)
        if meter_data:
            return jsonify({"success": True, "data": meter_data})
        return jsonify({"success": False, "error": "暂无历史抄表数据"})
    except Exception as e:
        return jsonify({"success": False, "error": f"服务器错误：{str(e)}"}), 500

@app.route('/api/calculate', methods=['POST'])
def api_calculate():
    """计算费用"""
    try:
        if not request.is_json:
            return jsonify({"success": False, "error": "请求格式必须为JSON"}), 400
        data = request.get_json()
        
        # 验证输入
        validate_result = validate_inputs(data)
        if not validate_result["valid"]:
            return jsonify({"success": False, "error": validate_result["error"]})
        
        # 保存/更新客户信息
        company_name = data.get('company_name', '')
        invoice_type = data.get('invoice_type', '增税')
        customer_id = add_or_update_customer(company_name, invoice_type)
        data['customer_id'] = customer_id
        
        # 计算费用
        calc_result = calculate_cost(data)
        if calc_result["success"]:
            add_calculation(calc_result["data"])
        return jsonify(calc_result)
    except Exception as e:
        return jsonify({"success": False, "error": f"服务器错误：{str(e)}"}), 500

@app.route('/api/export-excel', methods=['POST'])
def api_export_excel():
    """导出Excel"""
    try:
        export_result = export_to_excel()
        if export_result["success"]:
            return jsonify({
                "success": True,
                "filename": export_result["filename"],
                "download_url": f"/download/{export_result['filename']}",
                "total_basic_fee": export_result["total_basic_fee"],
                "total_over_fee": export_result["total_over_fee"],
                "total_all_fee": export_result["total_all_fee"]
            })
        else:
            return jsonify(export_result)
    except Exception as e:
        return jsonify({"success": False, "error": f"导出失败：{str(e)}"}), 500

@app.route('/download/<filename>')
def download_file(filename):
    """下载文件"""
    try:
        file_path = resource_path(filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            return jsonify({"success": False, "error": "文件不存在"}), 404
    except Exception as e:
        return jsonify({"success": False, "error": f"下载失败：{str(e)}"}), 500

@app.route('/api/clear-history', methods=['POST'])
def clear_history():
    """清空计算历史"""
    try:
        clear_calculations()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": f"清空失败：{str(e)}"}), 500

# ========================= 程序入口（修复启动逻辑） =========================
def run_server():
    """启动Flask服务（统一参数）"""
    try:
        # 确保templates和static文件夹存在
        os.makedirs(resource_path("templates"), exist_ok=True)
        os.makedirs(resource_path("static"), exist_ok=True)
        # 初始化数据库
        init_db()
        # 统一启动参数：禁用debug、自动重载，允许外部访问
        app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)
    except Exception as e:
        print(f"Flask服务启动失败：{e}")
        # 抛出异常让主线程捕获
        raise

def main():
    """主函数（打包/开发环境兼容）"""
    # 启动Flask服务（后台线程）
    flask_thread = threading.Thread(target=run_server, daemon=True)
    flask_thread.start()
    # 延长等待时间，确保服务完全启动
    time.sleep(3)
    
    # 尝试启动webview窗口，失败则用浏览器打开
    try:
        # 先安装webview依赖：pip install pywebview
        import webview
        # 正确创建并启动窗口
        window = webview.create_window(
            title="打印机费用计算系统",
            url="http://127.0.0.1:5000",
            width=1200,
            height=800,
            resizable=True
        )
        webview.start(debug=True)  # 开启debug便于排查窗口问题
    except ImportError:
        print("未安装pywebview，自动打开浏览器")
        webbrowser.open("http://127.0.0.1:5000")
        # 保持程序运行
        while True:
            time.sleep(1)
    except Exception as e:
        print(f"桌面窗口启动失败：{e}，自动打开浏览器")
        webbrowser.open("http://127.0.0.1:5000")
        while True:
            time.sleep(1)

if __name__ == "__main__":
    # 移除隐藏控制台的代码，便于调试
    # 安装缺失的依赖
    try:
        import webview
    except:
        print("请先安装pywebview：pip install pywebview")
    
    if getattr(sys, 'frozen', False):
        # 打包后的exe运行逻辑
        main()
    else:
        # 开发环境运行逻辑
        init_db()
        app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)